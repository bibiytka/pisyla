# vacancy_app.py
# Полное приложение для поиска вакансий HH.ru + SuperJob
# Для PythonAnywhere

from flask import Flask, request, Response, render_template_string, redirect, url_for, send_file
import requests
import json
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment

app = Flask(__name__)

@app.route('/favicon.ico')
def favicon():
    return redirect(url_for('static', filename='favicon.ico'), code=302)

# ============ ВШИТЫЕ CREDENTIALS ============
SUPERJOB_APP_ID = "4014"
SUPERJOB_SECRET_KEY = "v3.r.138979256.3b5b15795a107a49a55e7f4e5eed1857dfe78cde.dda83a292af5754f027da2f0c96152b9b34f0dee"

HH_CLIENT_ID = "OSG86UKP38OI11J3LN443EN7TU5NBIHVO9ACN312JO9O871KF3UNMRMHCJHF8AR5"
HH_CLIENT_SECRET = "LI36ECQCQNAHR9QV9U9MMPI1B2N9EK4I90QCA3ER02ADL6JQ5LPHBD2L6265M27R"
HH_REDIRECT_URI = "https://pisyla.onrender.com/hh_callback"

# ============ ПРОКСИ ДЛЯ SUPERJOB API ============
@app.route('/proxy')
def proxy():
    target_url = request.args.get('url')
    
    if not target_url:
        return {'error': 'URL parameter required'}, 400
    
    try:
        headers = {
            'User-Agent': 'VacancyParser/1.0'
        }
        
        # Передаем заголовок Authorization, если он есть в исходном запросе
        if 'Authorization' in request.headers:
            headers['Authorization'] = request.headers['Authorization']
        
        # Для SuperJob API также требуется X-Api-App-Id
        headers['X-Api-App-Id'] = SUPERJOB_SECRET_KEY
        
        response = requests.get(target_url, headers=headers, timeout=30)
        
        return Response(
            response.content,
            status=response.status_code,
            headers={
                'Access-Control-Allow-Origin': '*',
                'Content-Type': 'application/json; charset=utf-8'
            }
        )
    except Exception as e:
        return {'error': str(e)}, 500

# ============ OAUTH CALLBACK ============
@app.route('/callback')
def callback():
    return render_template_string(CALLBACK_HTML)

@app.route('/hh_callback')
def hh_callback():
    return render_template_string(CALLBACK_HTML)

# ============ OAUTH TOKEN EXCHANGE ============
@app.route('/get_sj_tokens', methods=['POST'])
def get_sj_tokens():
    data = request.get_json()
    code = data.get('code')
    redirect_uri = data.get('redirect_uri')
    client_id = data.get('client_id')
    client_secret = data.get('client_secret')

    if not all([code, redirect_uri, client_id, client_secret]):
        return {'error': 'Missing parameters'}, 400

    token_url = "https://api.superjob.ru/2.0/oauth2/access_token/"
    params = {
        'code': code,
        'redirect_uri': redirect_uri,
        'client_id': client_id,
        'client_secret': client_secret
    }

    try:
        response = requests.get(token_url, params=params, timeout=30)
        response.raise_for_status()
        return response.json()
    except requests.exceptions.RequestException as e:
        print(f"Error getting SJ tokens: {e}")
        return {'error': str(e)}, 500

@app.route('/refresh_sj_token', methods=['POST'])
def refresh_sj_token():
    data = request.get_json()
    refresh_token = data.get('refresh_token')
    client_id = data.get('client_id')
    client_secret = data.get('client_secret')

    if not all([refresh_token, client_id, client_secret]):
        return {'error': 'Missing parameters'}, 400

    refresh_url = "https://api.superjob.ru/2.0/oauth2/refresh_token/"
    params = {
        'refresh_token': refresh_token,
        'client_id': client_id,
        'client_secret': client_secret
    }

    try:
        response = requests.get(refresh_url, params=params, timeout=30)
        response.raise_for_status()
        return response.json()
    except requests.exceptions.RequestException as e:
        print(f"Error refreshing SJ token: {e}")
        return {'error': str(e)}, 500

@app.route('/get_hh_tokens', methods=['POST'])
def get_hh_tokens():
    data = request.get_json()
    code = data.get('code')
    redirect_uri = data.get('redirect_uri')

    if not all([code, redirect_uri]):
        return {'error': 'Missing parameters'}, 400

    token_url = "https://api.hh.ru/token"
    payload = {
        'grant_type': 'authorization_code',
        'client_id': HH_CLIENT_ID,
        'client_secret': HH_CLIENT_SECRET,
        'code': code,
        'redirect_uri': redirect_uri
    }
    headers = {'Content-Type': 'application/x-www-form-urlencoded'}

    try:
        response = requests.post(token_url, data=payload, headers=headers, timeout=30)
        response.raise_for_status()
        return response.json()
    except requests.exceptions.RequestException as e:
        print(f"Error getting HH tokens: {e}")
        return {'error': str(e)}, 500

@app.route('/refresh_hh_token', methods=['POST'])
def refresh_hh_token():
    data = request.get_json()
    refresh_token = data.get('refresh_token')

    if not refresh_token:
        return {'error': 'Missing parameters'}, 400

    refresh_url = "https://api.hh.ru/token"
    payload = {
        'grant_type': 'refresh_token',
        'client_id': HH_CLIENT_ID,
        'client_secret': HH_CLIENT_SECRET,
        'refresh_token': refresh_token
    }
    headers = {'Content-Type': 'application/x-www-form-urlencoded'}

    try:
        response = requests.post(refresh_url, data=payload, headers=headers, timeout=30)
        response.raise_for_status()
        return response.json()
    except requests.exceptions.RequestException as e:
        print(f"Error refreshing HH token: {e}")
        return {'error': str(e)}, 500

# ============ ГЛАВНАЯ СТРАНИЦА ============
@app.route('/')
def index():
    # Вставляем credentials в HTML через JavaScript
    html = MAIN_HTML.replace(
        '// CREDENTIALS_PLACEHOLDER',
        f'''
        // Автоматическая авторизация
        localStorage.setItem('sj_credentials', JSON.stringify({{
            client_id: '{SUPERJOB_APP_ID}',
            client_secret: '{SUPERJOB_SECRET_KEY}'
        }}));
        '''
    )
    return render_template_string(html)

# ============ ЭКСПОРТ В EXCEL ============
@app.route('/export_to_excel', methods=['POST'])
def export_to_excel():
    data = request.get_json()
    vacancies = data.get('vacancies', [])

    if not vacancies:
        return {'error': 'No vacancies provided'}, 400

    wb = Workbook()
    ws = wb.active
    ws.title = "Вакансии"

    # Заголовки
    headers = ["Должность", "Компания", "Зарплата", "Город", "Дата", "Телефон", "Email", "Имя контакта"]
    ws.append(headers)

    # Стили для заголовков
    header_font = Font(bold=True)
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    center_aligned_text = Alignment(horizontal="center")

    for col_num, cell in enumerate(ws[1], 1):
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = center_aligned_text
        # Устанавливаем ширину колонок
        if col_num == 1: ws.column_dimensions[chr(64 + col_num)].width = 40
        elif col_num == 2: ws.column_dimensions[chr(64 + col_num)].width = 30
        elif col_num == 3: ws.column_dimensions[chr(64 + col_num)].width = 20
        elif col_num == 4: ws.column_dimensions[chr(64 + col_num)].width = 20
        elif col_num == 5: ws.column_dimensions[chr(64 + col_num)].width = 15
        elif col_num == 6: ws.column_dimensions[chr(64 + col_num)].width = 25
        elif col_num == 7: ws.column_dimensions[chr(64 + col_num)].width = 30
        elif col_num == 8: ws.column_dimensions[chr(64 + col_num)].width = 25

    # Данные
    for vacancy in vacancies:
        ws.append([
            vacancy.get('title', ''),
            vacancy.get('company', ''),
            vacancy.get('salary', ''),
            vacancy.get('location', ''),
            vacancy.get('date', ''),
            vacancy.get('phone', ''),
            vacancy.get('email', ''),
            vacancy.get('contactName', '')
        ])
        # Применяем границы к ячейкам данных
        for cell in ws[ws.max_row]:
            cell.border = thin_border

    # Сохраняем книгу в байтовый поток
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    return send_file(
        excel_file,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='vacancies.xlsx'
    )

# ============ HTML CALLBACK СТРАНИЦЫ ============
CALLBACK_HTML = '''
<!DOCTYPE html>
<html lang="ru">
<head>
    <meta charset="UTF-8">
    <link rel="icon" href="/static/favicon.ico?v=2">
    <link rel="shortcut icon" href="/static/favicon.ico?v=2">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>SuperJob OAuth Callback</title>
    <style>
        body {
            font-family: 'Segoe UI', Arial, sans-serif;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            display: flex;
            align-items: center;
            justify-content: center;
            height: 100vh;
            margin: 0;
            text-align: center;
        }
        .container {
            background: rgba(255,255,255,0.1);
            backdrop-filter: blur(10px);
            padding: 40px;
            border-radius: 16px;
            box-shadow: 0 10px 40px rgba(0,0,0,0.3);
        }
        h1 { margin: 0 0 20px 0; }
        .spinner {
            border: 4px solid rgba(255,255,255,0.3);
            border-top: 4px solid white;
            border-radius: 50%;
            width: 50px;
            height: 50px;
            animation: spin 1s linear infinite;
            margin: 20px auto;
        }
        @keyframes spin {
            0% { transform: rotate(0deg); }
            100% { transform: rotate(360deg); }
        }
    </style>
</head>
<body>
    <div class="container">
        <h1>🔄 Обработка авторизации...</h1>
        <div class="spinner"></div>
        <p id="message">Пожалуйста, подождите</p>
    </div>
    <script>
        const urlParams = new URLSearchParams(window.location.search);
        const code = urlParams.get('code');
        const error = urlParams.get('error');
        const state = urlParams.get('state'); // Получаем параметр state
        const messageEl = document.getElementById('message');

        if (error) {
            messageEl.textContent = `❌ Ошибка: ${error}`;
            setTimeout(() => window.close(), 3000);
        } else if (code) {
            messageEl.textContent = `✅ Код получен: ${code.substring(0, 10)}...`;
            if (window.opener) {
                if (state === 'hh_auth') {
                    window.opener.postMessage({ type: 'hh_auth_success', code }, '*');
                } else if (state === 'superjob_auth') {
                    window.opener.postMessage({ type: 'sj_auth_success', code }, '*');
                }
                messageEl.textContent = '✅ Готово! Можно закрыть это окно.';
                setTimeout(() => window.close(), 1000);
            } else {
                messageEl.textContent = '✅ Вернитесь на главную страницу';
            }
        } else {
            messageEl.textContent = '⚠️ Не получен код авторизации';
        }
    </script>
</body>
</html>
'''

# ============ ГЛАВНЫЙ HTML ============
MAIN_HTML = '''
<!DOCTYPE html>
<html lang="ru">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Вакансии - HH.ru и SuperJob</title>

    <!-- ВСТАВИТЬ ЭТИ 2 СТРОКИ -->
    <link rel="icon" href="/static/favicon.ico?v=2">
    <link rel="shortcut icon" href="/static/favicon.ico?v=2">

    <style>
        * { box-sizing: border-box; }
        body {
            font-family: 'Segoe UI', Arial, sans-serif;
            background: #f0f4f8;
            color: #333;
            margin: 0;
            padding: 20px;
            line-height: 1.6;
        }
        .container {
            max-width: 1400px;
            margin: 0 auto;
            background: white;
            padding: 30px;
            border-radius: 12px;
            box-shadow: 0 6px 20px rgba(0,0,0,0.1);
        }
        h1 {
            text-align: center;
            color: #2c3e50;
            margin-bottom: 10px;
            font-size: 28px;
        }
        .subtitle {
            text-align: center;
            color: #7f8c8d;
            font-size: 16px;
            margin-bottom: 30px;
        }

        .auth-status {
            background: linear-gradient(135deg, #ff8800 0%, #ff6b35 100%);
            color: white;
            padding: 20px;
            border-radius: 8px;
            margin: 20px 0;
            display: flex;
            justify-content: space-between;
            align-items: center;
        }
        .auth-status.logged-out {
            background: linear-gradient(135deg, #95a5a6 0%, #7f8c8d 100%);
        }
        .auth-info { flex: 1; }
        .auth-info h3 { margin: 0 0 5px 0; }
        .auth-info p { margin: 0; opacity: 0.9; font-size: 14px; }
        .auth-actions button {
            background: rgba(255,255,255,0.2);
            color: white;
            border: 2px solid white;
            padding: 10px 20px;
            border-radius: 6px;
            cursor: pointer;
            font-weight: bold;
        }
        .auth-actions button:hover {
            background: white;
            color: #ff8800;
        }

        .source-selector {
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            padding: 25px;
            border-radius: 8px;
            margin: 20px 0;
            color: white;
        }
        .source-selector h3 {
            margin: 0 0 15px 0;
            text-align: center;
        }
        .source-options {
            display: flex;
            justify-content: center;
            gap: 20px;
            flex-wrap: wrap;
        }
        .source-option {
            background: rgba(255, 255, 255, 0.15);
            backdrop-filter: blur(10px);
            border: 2px solid rgba(255, 255, 255, 0.3);
            border-radius: 8px;
            padding: 20px 30px;
            cursor: pointer;
            transition: all 0.3s;
            min-width: 200px;
            text-align: center;
        }
        .source-option:hover {
            background: rgba(255, 255, 255, 0.25);
            transform: translateY(-3px);
        }
        .source-option.active {
            background: white;
            color: #667eea;
            border-color: white;
            box-shadow: 0 8px 20px rgba(0,0,0,0.2);
        }
        .source-option input[type="radio"] { display: none; }
        .source-option .source-logo {
            font-size: 32px;
            margin-bottom: 10px;
        }
        .source-option .source-name {
            font-weight: bold;
            font-size: 18px;
        }

        .filters-box {
            background: #f8f9fa;
            border-radius: 8px;
            padding: 25px;
            margin: 20px 0;
            border: 1px solid #e9ecef;
            position: relative; /* Добавлено для абсолютного позиционирования дочерних элементов */
        }
        .filter-row {
            display: flex;
            flex-wrap: wrap;
            gap: 20px;
            margin-bottom: 20px;
            align-items: flex-start;
        }
        .filter-group {
            flex: 1;
            min-width: 250px;
        }
        .filter-group label {
            display: block;
            font-weight: bold;
            color: #2c3e50;
            margin-bottom: 8px;
            font-size: 14px;
        }
        .filter-group small {
            color: #7f8c8d;
            font-weight: normal;
            display: block;
            margin-top: 4px;
        }
        input[type="text"] {
            width: 100%;
            padding: 12px;
            border: 2px solid #ddd;
            border-radius: 6px;
            font-size: 15px;
            transition: border 0.3s;
            font-family: inherit;
        }
        input[type="text"]:focus {
            outline: none;
            border-color: #1a73e8;
        }

        .city-selector { position: relative; }
        .city-dropdown {
            width: 100%;
            padding: 12px;
            border: 2px solid #ddd;
            border-radius: 6px;
            background: white;
            cursor: pointer;
            font-size: 15px;
            display: flex;
            justify-content: space-between;
            align-items: center;
        }
        .city-dropdown:hover { border-color: #1a73e8; }
        .city-options {
            display: none;
            position: absolute;
            top: 100%;
            left: 0;
            right: 0;
            background: white;
            border: 2px solid #1a73e8;
            border-radius: 6px;
            margin-top: 5px;
            max-height: 300px;
            overflow-y: auto;
            z-index: 1000;
            box-shadow: 0 4px 12px rgba(0,0,0,0.15);
        }
        .city-options.active { display: block; }
        .city-option {
            padding: 10px 15px;
            cursor: pointer;
            transition: background 0.2s;
            border-bottom: 1px solid #f0f0f0;
        }
        .city-option:hover { background: #f0f8ff; }
        .city-option input[type="checkbox"] { margin-right: 10px; }

        .checkbox-group {
            display: flex;
            flex-wrap: wrap;
            gap: 15px;
            margin-top: 15px;
        }
        .checkbox-group label {
            display: flex;
            align-items: center;
            cursor: pointer;
            font-weight: normal;
        }
        .checkbox-group input[type="checkbox"] {
            margin-right: 6px;
            cursor: pointer;
        }

        .exclusion-tags {
            display: flex;
            flex-wrap: wrap;
            gap: 8px;
            margin-top: 10px;
            min-height: 40px;
            padding: 8px;
            border: 2px dashed #ddd;
            border-radius: 6px;
        }
        .exclusion-tag {
            background: #e74c3c;
            color: white;
            padding: 6px 12px;
            border-radius: 20px;
            font-size: 14px;
            display: inline-flex;
            align-items: center;
            gap: 8px;
        }
        .exclusion-tag .remove-tag {
            cursor: pointer;
            font-weight: bold;
            font-size: 16px;
        }
        .add-exclusion-btn {
            background: #95a5a6;
            color: white;
            border: none;
            padding: 8px 16px;
            border-radius: 6px;
            cursor: pointer;
            font-size: 14px;
            margin-top: 8px;
        }

        button {
            padding: 14px 24px;
            font-size: 16px;
            font-weight: bold;
            border: none;
            border-radius: 6px;
            cursor: pointer;
            transition: all 0.3s;
        }
        button:hover:not(:disabled) {
            transform: translateY(-2px);
            box-shadow: 0 4px 10px rgba(0,0,0,0.15);
        }
        button:disabled {
            opacity: 0.5;
            cursor: not-allowed;
        }
        .btn-search {
            background: #1a73e8;
            color: white;
            width: 100%;
            font-size: 18px;
            padding: 16px;
        }
        .btn-hh { background: #00a0e1; color: white; }
        .btn-avito { background: #005bb9; color: white; }
        .btn-sj { background: #ff8800; color: white; }
        .btn-zp { background: #4caf50; color: white; }

        .external-links {
            text-align: center;
            margin: 20px 0;
            padding: 15px;
            background: #fff9e6;
            border-radius: 8px;
            border: 1px solid #ffe6a0;
        }
        .external-links button { margin: 5px; }

        .stats {
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 20px;
            border-radius: 8px;
            margin: 20px 0;
            text-align: center;
            box-shadow: 0 4px 12px rgba(102, 126, 234, 0.3);
        }
        .stats-grid {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(150px, 1fr));
            gap: 15px;
            margin-top: 15px;
        }
        .stat-item {
            background: rgba(255,255,255,0.1);
            padding: 15px;
            border-radius: 6px;
            backdrop-filter: blur(10px);
        }
        .stat-number {
            font-size: 32px;
            font-weight: bold;
            display: block;
        }
        .stat-label {
            font-size: 14px;
            opacity: 0.9;
        }
        .stats .source-badge {
            display: inline-block;
            padding: 5px 15px;
            background: rgba(255,255,255,0.2);
            border-radius: 20px;
            margin-left: 10px;
            font-size: 14px;
        }

        .results { margin-top: 40px; }
        table {
            width: 100%;
            border-collapse: collapse;
            margin: 20px 0;
            font-size: 14px;
            box-shadow: 0 1px 5px rgba(0,0,0,0.1);
        }
        th, td {
            border: 1px solid #ddd;
            padding: 14px;
            text-align: left;
            vertical-align: top;
        }
        th {
            background: #34495e;
            color: white;
            font-weight: bold;
            position: sticky;
            top: 0;
            z-index: 10;
        }
        tr:nth-child(even) { background: #f9f9f9; }
        tr:hover {
            background: #f0f8ff;
            cursor: pointer;
        }
        .vacancy-title {
            font-weight: bold;
            color: #2c3e50;
        }
        .company-name {
            color: #7f8c8d;
            font-weight: bold;
        }
        .salary {
            color: #27ae60;
            font-weight: bold;
        }
        .location {
            color: #3498db;
            font-size: 13px;
        }
        .date {
            color: #95a5a6;
            font-size: 12px;
        }

        .loader {
            text-align: center;
            padding: 40px;
            font-size: 16px;
            color: #7f8c8d;
        }
        .spinner {
            border: 4px solid #f3f3f3;
            border-top: 4px solid #1a73e8;
            border-radius: 50%;
            width: 40px;
            height: 40px;
            animation: spin 1s linear infinite;
            margin: 20px auto;
        }
        @keyframes spin {
            0% { transform: rotate(0deg); }
            100% { transform: rotate(360deg); }
        }

        .no-results {
            text-align: center;
            padding: 60px;
            color: #e74c3c;
            font-size: 18px;
        }

        .scroll-top {
            position: fixed;
            bottom: 30px;
            right: 30px;
            background: #1a73e8;
            color: white;
            border: none;
            border-radius: 50%;
            width: 50px;
            height: 50px;
            font-size: 24px;
            cursor: pointer;
            display: none;
            box-shadow: 0 4px 12px rgba(0,0,0,0.2);
            z-index: 1000;
        }
        .scroll-top.visible { display: block; }

        .badge-new {
            background: #e74c3c;
            color: white;
            padding: 3px 8px;
            border-radius: 4px;
            font-size: 11px;
            margin-left: 8px;
            font-weight: bold;
        }

        .badge-source {
            background: #95a5a6;
            color: white;
            padding: 2px 6px;
            border-radius: 3px;
            font-size: 10px;
            margin-left: 6px;
            font-weight: normal;
        }
        .badge-source.hh { background: #00a0e1; }
        .badge-source.sj { background: #ff8800; }

        .contacts {
            font-size: 13px;
            line-height: 1.4;
        }
        .contact-phone {
            color: #27ae60;
            font-weight: bold;
            white-space: nowrap;
        }
        .contact-email {
            color: #3498db;
            font-size: 12px;
        }
        .contact-name {
            color: #7f8c8d;
            font-size: 11px;
            font-style: italic;
        }
        .no-contacts {
            color: #bdc3c7;
            font-style: italic;
            font-size: 12px;
        }

        @media (max-width: 768px) {
            body { padding: 10px; }
            .container { padding: 15px; }
            h1 { font-size: 20px; }
        }
    </style>
</head>
<body>
    <div class="container">
        <h1>🏭 Вакансии линейного персонала</h1>
        <p class="subtitle">Грузчики • Кладовщики • Комплектовщики • Упаковщики • Разнорабочие</p>

        <div class="auth-status logged-out" id="sjAuthStatus" style="display: none;">
            <div class="auth-info">
                <h3 id="sjAuthTitle">🟠 SuperJob: Не авторизован</h3>
                <p id="sjAuthText">Для поиска вакансий SuperJob требуется авторизация.</p>
            </div>
            <div class="auth-actions">
                <button onclick="authorizeSuperJob()" id="btnSjAuthorize">Войти через SuperJob</button>
                <button onclick="logout('superjob')" id="btnSjLogout" style="display: none;">Выйти</button>
            </div>
        </div>

        <div class="auth-status logged-out" id="hhAuthStatus">
            <div class="auth-info">
                <h3 id="hhAuthTitle">🔵 HH.ru: Не авторизован</h3>
                <p id="hhAuthText">Для получения контактов HH.ru требуется авторизация.</p>
            </div>
            <div class="auth-actions">
                <button onclick="authorizeHH()" id="btnHhAuthorize">Войти через HH.ru</button>
                <button onclick="logout('hh')" id="btnHhLogout" style="display: none;">Выйти</button>
            </div>
        </div>

        <div class="source-selector">
            <h3>🎯 Выберите источник вакансий</h3>
            <div class="source-options">
                <label class="source-option active" for="sourceHH">
                    <input type="radio" name="source" id="sourceHH" value="hh" checked onchange="switchSource('hh')">
                    <div class="source-logo">🔵</div>
                    <div class="source-name">HH.ru</div>
                    <small>Без авторизации</small>
                </label>
                <label class="source-option" for="sourceSJ">
                    <input type="radio" name="source" id="sourceSJ" value="superjob" onchange="switchSource('superjob')">
                    <div class="source-logo">🟠</div>
                    <div class="source-name">SuperJob</div>
                    <small>Готов к работе</small>
                </label>
            </div>
        </div>

        <div class="filters-box">
            <h3 style="margin-top: 0; color: #2c3e50;">🔍 Настройте поиск</h3>
            
            <div class="filter-row">
                <div class="filter-group">
                    <label for="query">Должность:</label>
                    <input type="text" id="query" placeholder="Например: грузчик" value="склад">
                </div>

                <div class="filter-group">
                    <label>Город(а):</label>
                    <div class="city-selector">
                        <div class="city-dropdown" onclick="toggleCityDropdown()">
                            <span id="selectedCitiesText">Санкт-Петербург</span>
                            <span>▼</span>
                        </div>
                        <div class="city-options" id="cityOptions"></div>
                    </div>
                </div>
            </div>

            <div id="exclusionBlock" class="filter-row">
                <div class="filter-group" style="flex: 1 1 100%;">
                    <label for="exclusionInput">
                        Слова-исключения (HH.ru):
                        <small>Вакансии с этими словами будут скрыты</small>
                    </label>
                    <input type="text" id="exclusionInput" placeholder="Введите слово">
                    <button class="add-exclusion-btn" onclick="addExclusion()">➕ Добавить</button>
                    <div class="exclusion-tags" id="exclusionTags">
                        <small style="color: #7f8c8d;">Нет исключений</small>
                    </div>
                </div>
            </div>

            <div class="checkbox-group">
                <label>
                    <input type="checkbox" id="freshOnly" checked> За последние 30 дней
                </label>
                <label>
                    <input type="checkbox" id="lastTwoDays"> За последние 2 дня
                </label>
                <label>
                    <input type="checkbox" id="withSalary"> С зарплатой
                </label>
                <label>
                    <input type="checkbox" id="noExp" checked> Без опыта
                </label>
                <label>
                    <input type="checkbox" id="fullTime" checked> Полная занятость
                </label>
                <label>
                    <input type="checkbox" id="oneVacancyPerCompany"> Одна вакансия от компании
                </label>
                <label>
                    <input type="checkbox" id="filterBySuffix"> 
                    Фильтр по суффиксам
                    <small title="Исключает вакансии с другими профессиями на 'ик', 'ор', 'ий'">
                        ❓ Только целевая профессия (ик/ор/ий)
                    </small>
                </label>
                <label>
                    <input type="checkbox" id="onlyWithContacts"> Только с контактами
                </label>
            </div>

            <button onclick="startNewSearch()" class="btn-search" style="margin-bottom: 10px;">🔎 Найти вакансии</button>
        </div>

        <button onclick="exportToExcel()" class="btn-search" style="background: #28a745; position: fixed; bottom: 90px; right: 20px; z-index: 9999; width: auto;">📊 Экспорт в Excel</button>

        <div class="stats" id="stats" style="display: none;">
            <h3 style="margin: 0 0 10px 0;">
                📊 Статистика
                <span class="source-badge" id="currentSourceBadge">HH.ru</span>
            </h3>
            <div class="stats-grid">
                <div class="stat-item">
                    <span class="stat-number" id="statTotal">0</span>
                    <span class="stat-label">Всего найдено</span>
                </div>
                <div class="stat-item">
                    <span class="stat-number" id="statLoaded">0</span>
                    <span class="stat-label">Загружено</span>
                </div>
                <div class="stat-item">
                    <span class="stat-number" id="statWithSalary">0</span>
                    <span class="stat-label">С зарплатой</span>
                </div>
                <div class="stat-item">
                    <span class="stat-number" id="statExcluded">0</span>
                    <span class="stat-label">Исключено</span>
                </div>
                <div class="stat-item" id="statSuffixExcluded" style="display:none;">
                    <span class="stat-number" id="statSuffixCount">0</span>
                    <span class="stat-label">
                        Исключено по суффиксам
                        <span id="suffixDetails" style="font-size:10px; display:block;"></span>
                    </span>
                </div>
            </div>
        </div>

        <div class="external-links">
            <p style="margin: 0 0 10px 0; font-weight: bold; color: #856404;">🌐 Открыть на сайтах:</p>
            <button onclick="go('hh')" class="btn-hh">HH.ru</button>
            <button onclick="go('avito')" class="btn-avito">Avito</button>
            <button onclick="go('superjob')" class="btn-sj">SuperJob</button>
            <button onclick="go('zarplata')" class="btn-zp">Zarplata.ru</button>
        </div>

        <div class="results">
            <h2>📋 Результаты</h2>
            <table id="vacancyTable">
                <thead>
                    <tr>
                        <th style="width: 25%;">Должность</th>
                        <th style="width: 20%;">Компания</th>
                        <th style="width: 12%;">Зарплата</th>
                        <th style="width: 12%;">Город</th>
                        <th style="width: 8%;">Дата</th>
                        <th style="width: 23%;">Контакты</th>
                    </tr>
                </thead>
                <tbody id="vacancyTableBody">
                    <tr><td colspan="6" class="no-results">Нажмите "Найти вакансии"</td></tr>
                </tbody>
            </table>

            <div id="loadingIndicator" class="loader" style="display: none;">
                <div class="spinner"></div>
                <p>Загружаем...</p>
            </div>
        </div>
    </div>

    <button class="scroll-top" id="scrollTopBtn" onclick="scrollToTop()">↑</button>

    <script>
        // CREDENTIALS_PLACEHOLDER
        
        function delay(ms) {
            return new Promise(resolve => setTimeout(resolve, ms));
        }

        const API_BASE = 'https://api.superjob.ru/2.0';
        const SUPERJOB_CLIENT_ID = '4014'; // Ваш client_id
        const SUPERJOB_SECRET_KEY = 'v3.r.138979256.3b5b15795a107a49a55e7f4e5eed1857dfe78cde.dda83a292af5754f027da2f0c96152b9b34f0dee'; // Ваш secret_key
        const SUPERJOB_REDIRECT_URI = 'https://pisyla.onrender.com/callback';

        const HH_CLIENT_ID = 'OSG86UKP38OI11J3LN443EN7TU5NBIHVO9ACN312JO9O871KF3UNMRMHCJHF8AR5';
        const HH_CLIENT_SECRET = 'LI36ECQCQNAHR9QV9U9MMPI1B2N9EK4I90QCA3ER02ADL6JQ5LPHBD2L6265M27R';
        const HH_REDIRECT_URI = 'https://pisyla.onrender.com/hh_callback';

        let sjAccessToken = localStorage.getItem('sj_access_token');
        let sjRefreshToken = localStorage.getItem('sj_refresh_token');
        let sjExpiresIn = localStorage.getItem('sj_expires_in');
        let sjAuthWindow = null;

        let hhAccessToken = localStorage.getItem('hh_access_token');
        let hhRefreshToken = localStorage.getItem('hh_refresh_token');
        let hhExpiresIn = localStorage.getItem('hh_expires_in');
        let hhAuthWindow = null;

        let currentPage = 0;
        let isLoading = false;
        let hasMore = true;
        let totalFound = 0;
        let loadedCount = 0;
        let withSalaryCount = 0;
        let excludedCount = 0;
        let excludedBySuffixDetails = { ик: 0, ор: 0, ий: 0 }; // Добавляем переменную для статистики по суффиксам
        let vacanciesPerSecond = 0; // Новая переменная для скорости загрузки
        let currentQuery = '';
        let currentSource = 'hh';
        let exclusionWords = [];
        let allLoadedVacancies = []; // Добавляем массив для хранения всех загруженных вакансий

        const citiesHH = {
            1: 'Москва', 2: 'Санкт-Петербург', 3: 'Екатеринбург',
            4: 'Новосибирск', 66: 'Нижний Новгород', 88: 'Казань',
            78: 'Самара', 76: 'Краснодар', 54: 'Ростов-на-Дону', 113: 'Вся Россия'
        };

        const citiesSJ = {
            4: 'Москва', 14: 'Санкт-Петербург', 33: 'Екатеринбург',
            13: 'Новосибирск', 12: 'Нижний Новгород', 55: 'Казань',
            5: 'Самара', 25: 'Краснодар', 73: 'Ростов-на-Дону', 0: 'Вся Россия'
        };

        let currentCities = [];

        function logout(source) {
            if (source === 'superjob') {
                if (confirm('Вы уверены, что хотите выйти из SuperJob?')) {
                    localStorage.removeItem('sj_access_token');
                    localStorage.removeItem('sj_refresh_token');
                    localStorage.removeItem('sj_expires_in');
                    sjAccessToken = null;
                    sjRefreshToken = null;
                    sjExpiresIn = null;
                    updateAuthStatus('superjob');
                    alert('Вы вышли из SuperJob. Перезагрузите страницу для повторной авторизации.');
                }
            } else if (source === 'hh') {
                if (confirm('Вы уверены, что хотите выйти из HH.ru?')) {
                    localStorage.removeItem('hh_access_token');
                    localStorage.removeItem('hh_refresh_token');
                    localStorage.removeItem('hh_expires_in');
                    hhAccessToken = null;
                    hhRefreshToken = null;
                    hhExpiresIn = null;
                    updateAuthStatus('hh');
                    alert('Вы вышли из HH.ru. Перезагрузите страницу для повторной авторизации.');
                }
            }
        }

        function authorizeSuperJob() {
            const authUrl = `https://www.superjob.ru/authorize/?client_id=${SUPERJOB_CLIENT_ID}&redirect_uri=${encodeURIComponent(SUPERJOB_REDIRECT_URI)}&state=superjob_auth`;
            sjAuthWindow = window.open(authUrl, '_blank', 'width=600,height=700');
        }

        function authorizeHH() {
            const authUrl = `https://hh.ru/oauth/authorize?response_type=code&client_id=${HH_CLIENT_ID}&redirect_uri=${encodeURIComponent(HH_REDIRECT_URI)}&state=hh_auth`;
            hhAuthWindow = window.open(authUrl, '_blank', 'width=600,height=700');
        }

        async function getSjTokens(code) {
            try {
                const response = await fetch('/get_sj_tokens', {
                    method: 'POST',
                    headers: { 'Content-Type': 'application/json' },
                    body: JSON.stringify({
                        code: code,
                        redirect_uri: SUPERJOB_REDIRECT_URI,
                        client_id: SUPERJOB_CLIENT_ID,
                        client_secret: SUPERJOB_SECRET_KEY
                    })
                });
                if (!response.ok) throw new Error(`HTTP ${response.status}`);
                const data = await response.json();
                
                sjAccessToken = data.access_token;
                sjRefreshToken = data.refresh_token;
                sjExpiresIn = Date.now() + (data.expires_in * 1000); // Время истечения в миллисекундах
                
                localStorage.setItem('sj_access_token', sjAccessToken);
                localStorage.setItem('sj_refresh_token', sjRefreshToken);
                localStorage.setItem('sj_expires_in', sjExpiresIn);
                
                updateAuthStatus('superjob');
                return true;
            } catch (error) {
                console.error('Ошибка при получении токенов SuperJob:', error);
                alert('Ошибка авторизации SuperJob. Пожалуйста, попробуйте снова.');
                return false;
            }
        }

        async function refreshSjToken() {
            if (!sjRefreshToken) return false;
            try {
                const response = await fetch('/refresh_sj_token', {
                    method: 'POST',
                    headers: { 'Content-Type': 'application/json' },
                    body: JSON.stringify({
                        refresh_token: sjRefreshToken,
                        client_id: SUPERJOB_CLIENT_ID,
                        client_secret: SUPERJOB_SECRET_KEY
                    })
                });
                if (!response.ok) throw new Error(`HTTP ${response.status}`);
                const data = await response.json();

                sjAccessToken = data.access_token;
                sjRefreshToken = data.refresh_token;
                sjExpiresIn = Date.now() + (data.expires_in * 1000);

                localStorage.setItem('sj_access_token', sjAccessToken);
                localStorage.setItem('sj_refresh_token', sjRefreshToken);
                localStorage.setItem('sj_expires_in', sjExpiresIn);
                
                updateAuthStatus('superjob');
                return true;
            } catch (error) {
                console.error('Ошибка при обновлении токена SuperJob:', error);
                logout('superjob'); // Если не удалось обновить, выходим
                return false;
            }
        }

        async function getHhTokens(code) {
            try {
                const response = await fetch('/get_hh_tokens', {
                    method: 'POST',
                    headers: { 'Content-Type': 'application/json' },
                    body: JSON.stringify({
                        code: code,
                        redirect_uri: HH_REDIRECT_URI
                    })
                });
                if (!response.ok) throw new Error(`HTTP ${response.status}`);
                const data = await response.json();
                
                hhAccessToken = data.access_token;
                hhRefreshToken = data.refresh_token;
                hhExpiresIn = Date.now() + (data.expires_in * 1000);
                
                localStorage.setItem('hh_access_token', hhAccessToken);
                localStorage.setItem('hh_refresh_token', hhRefreshToken);
                localStorage.setItem('hh_expires_in', hhExpiresIn);
                
                updateAuthStatus('hh');
                return true;
            } catch (error) {
                console.error('Ошибка при получении токенов HH.ru:', error);
                alert('Ошибка авторизации HH.ru. Пожалуйста, попробуйте снова.');
                return false;
            }
        }

        async function refreshHhToken() {
            if (!hhRefreshToken) return false;
            try {
                const response = await fetch('/refresh_hh_token', {
                    method: 'POST',
                    headers: { 'Content-Type': 'application/json' },
                    body: JSON.stringify({
                        refresh_token: hhRefreshToken
                    })
                });
                if (!response.ok) throw new Error(`HTTP ${response.status}`);
                const data = await response.json();

                hhAccessToken = data.access_token;
                hhRefreshToken = data.refresh_token;
                hhExpiresIn = Date.now() + (data.expires_in * 1000);

                localStorage.setItem('hh_access_token', hhAccessToken);
                localStorage.setItem('hh_refresh_token', hhRefreshToken);
                localStorage.setItem('hh_expires_in', hhExpiresIn);
                
                updateAuthStatus('hh');
                return true;
            } catch (error) {
                console.error('Ошибка при обновлении токена HH.ru:', error);
                logout('hh'); // Если не удалось обновить, выходим
                return false;
            }
        }

        function updateAuthStatus(source = 'all') {
            if (source === 'superjob' || source === 'all') {
                const authStatusDiv = document.getElementById('sjAuthStatus');
                const authTitle = document.getElementById('sjAuthTitle');
                const authText = document.getElementById('sjAuthText');
                const btnAuthorize = document.getElementById('btnSjAuthorize');
                const btnLogout = document.getElementById('btnSjLogout');

                if (sjAccessToken && sjExpiresIn && Date.now() < parseInt(sjExpiresIn)) {
                    authStatusDiv.classList.remove('logged-out');
                    authStatusDiv.style.background = 'linear-gradient(135deg, #2ecc71 0%, #27ae60 100%)'; // Зеленый
                    authTitle.textContent = '✅ SuperJob: Авторизован';
                    authText.textContent = `Токен действителен до: ${new Date(parseInt(sjExpiresIn)).toLocaleTimeString()} ${new Date(parseInt(sjExpiresIn)).toLocaleDateString()}`;
                    btnAuthorize.style.display = 'none';
                    btnLogout.style.display = 'block';
                } else {
                    authStatusDiv.classList.add('logged-out');
                    authStatusDiv.style.background = 'linear-gradient(135deg, #95a5a6 0%, #7f8c8d 100%)'; // Серый
                    authTitle.textContent = '🟠 SuperJob: Не авторизован';
                    authText.textContent = 'Для поиска вакансий SuperJob требуется авторизация.';
                    btnAuthorize.style.display = 'block';
                    btnLogout.style.display = 'none';
                }
            }

            if (source === 'hh' || source === 'all') {
                const authStatusDiv = document.getElementById('hhAuthStatus');
                const authTitle = document.getElementById('hhAuthTitle');
                const authText = document.getElementById('hhAuthText');
                const btnAuthorize = document.getElementById('btnHhAuthorize');
                const btnLogout = document.getElementById('btnHhLogout');

                if (hhAccessToken && hhExpiresIn && Date.now() < parseInt(hhExpiresIn)) {
                    authStatusDiv.classList.remove('logged-out');
                    authStatusDiv.style.background = 'linear-gradient(135deg, #2ecc71 0%, #27ae60 100%)'; // Зеленый
                    authTitle.textContent = '✅ HH.ru: Авторизован';
                    authText.textContent = `Токен действителен до: ${new Date(parseInt(hhExpiresIn)).toLocaleTimeString()} ${new Date(parseInt(hhExpiresIn)).toLocaleDateString()}`;
                    btnAuthorize.style.display = 'none';
                    btnLogout.style.display = 'block';
                } else {
                    authStatusDiv.classList.add('logged-out');
                    authStatusDiv.style.background = 'linear-gradient(135deg, #95a5a6 0%, #7f8c8d 100%)'; // Серый
                    authTitle.textContent = '🔵 HH.ru: Не авторизован';
                    authText.textContent = 'Для получения контактов HH.ru требуется авторизация.';
                    btnAuthorize.style.display = 'block';
                    btnLogout.style.display = 'none';
                }
            }
        }

        window.addEventListener('message', async (event) => {
            if (event.origin !== window.location.origin) return;
            if (event.data.type === 'sj_auth_success' && event.data.code) {
                if (sjAuthWindow) sjAuthWindow.close();
                await getSjTokens(event.data.code);
            } else if (event.data.type === 'hh_auth_success' && event.data.code) {
                if (hhAuthWindow) hhAuthWindow.close();
                await getHhTokens(event.data.code);
            }
        });

        function getDateDaysAgo(days) {
            const date = new Date();
            date.setDate(date.getDate() - days);
            return date.toISOString().split('T')[0];
        }

        function getUnixTimeDaysAgo(days) {
            const date = new Date();
            date.setDate(date.getDate() - days);
            return Math.floor(date.getTime() / 1000);
        }

        function formatDate(dateString) {
            const date = new Date(dateString);
            const now = new Date();
            const diffDays = Math.ceil(Math.abs(now - date) / (1000 * 60 * 60 * 24));
            
            if (diffDays === 1) return 'Сегодня';
            if (diffDays === 2) return 'Вчера';
            if (diffDays <= 7) return `${diffDays - 1} дн. назад`;
            
            return date.toLocaleDateString('ru-RU', { day: 'numeric', month: 'short' });
        }

        function formatDateFromUnix(unixtime) {
            return formatDate(new Date(unixtime * 1000).toISOString());
        }

        function isNew(dateString) {
            const diffHours = Math.abs(new Date() - new Date(dateString)) / (1000 * 60 * 60);
            return diffHours < 24;
        }

        function isNewFromUnix(unixtime) {
            return isNew(new Date(unixtime * 1000).toISOString());
        }

        function switchSource(source) {
            currentSource = source;
            
            document.querySelectorAll('.source-option').forEach(opt => opt.classList.remove('active'));
            document.querySelector(`label[for="source${source === 'hh' ? 'HH' : 'SJ'}"]`).classList.add('active');
            
            document.getElementById('exclusionBlock').style.display = source === 'hh' ? 'block' : 'none';

            // Control visibility of auth status divs
            document.getElementById('hhAuthStatus').style.display = source === 'hh' ? 'flex' : 'none';
            document.getElementById('sjAuthStatus').style.display = source === 'superjob' ? 'flex' : 'none';
            
            // Set default city for the new source
            const defaultId = currentSource === 'hh' ? 2 : 4;
            currentCities = [defaultId]; // Reset to default for the new source
            
            loadCities();
        }

        function loadCities() {
            const cities = currentSource === 'hh' ? citiesHH : citiesSJ;
            const container = document.getElementById('cityOptions');
            container.innerHTML = '';
            
            for (const [id, name] of Object.entries(cities)) {
                const cityId = parseInt(id);
                const isChecked = currentCities.includes(cityId); // Check if city is in currentCities
                
                const div = document.createElement('div');
                div.className = 'city-option';
                div.innerHTML = `
                    <label>
                        <input type="checkbox" value="${id}" ${isChecked ? 'checked' : ''} onchange="updateCitySelection()">
                        ${name}
                    </label>
                `;
                container.appendChild(div);
            }
            
            updateCitySelection();
        }

        function toggleCityDropdown() {
            document.getElementById('cityOptions').classList.toggle('active');
        }

        document.addEventListener('click', (e) => {
            if (!e.target.closest('.city-selector')) {
                document.getElementById('cityOptions').classList.remove('active');
            }
        });

        function updateCitySelection() {
            const checkboxes = document.querySelectorAll('#cityOptions input:checked');
            currentCities = Array.from(checkboxes).map(cb => parseInt(cb.value));
            
            const cities = currentSource === 'hh' ? citiesHH : citiesSJ;
            const text = currentCities.length === 0 ? 'Выберите город' :
                         currentCities.length === 1 ? cities[currentCities[0]] :
                         `Выбрано: ${currentCities.length}`;
            document.getElementById('selectedCitiesText').textContent = text;
        }

        function addExclusion() {
            const word = document.getElementById('exclusionInput').value.trim().toLowerCase();
            if (word && !exclusionWords.includes(word)) {
                exclusionWords.push(word);
                updateExclusionTags();
                document.getElementById('exclusionInput').value = '';
            }
        }

        function removeExclusion(word) {
            exclusionWords = exclusionWords.filter(w => w !== word);
            updateExclusionTags();
        }

        function updateExclusionTags() {
            const container = document.getElementById('exclusionTags');
            container.innerHTML = exclusionWords.length === 0 
                ? '<small style="color: #7f8c8d;">Нет исключений</small>'
                : exclusionWords.map(w => `
                    <div class="exclusion-tag">
                        <span>${w}</span>
                        <span class="remove-tag" onclick="removeExclusion('${w}')">×</span>
                    </div>
                `).join('');
        }

        function isExcluded(text) {
            if (!text || exclusionWords.length === 0) return false;
            return exclusionWords.some(word => text.toLowerCase().includes(word));
        }

        function shouldExcludeBySuffix(vacancyTitle, query) {
            const filterBySuffixChecked = document.getElementById('filterBySuffix')?.checked;
            if (!filterBySuffixChecked) return false;

            const queryLower = query.toLowerCase().trim();
            const titleLower = vacancyTitle.toLowerCase();

            // Разбиваем на слова (дефис теперь является разделителем)
            const queryWords = queryLower.split(/[^\wа-яё]+/ui).filter(w => w.length > 0);
            const titleWords = titleLower.split(/[^\wа-яё]+/ui).filter(w => w.length > 0);
            
            // Суффиксы для проверки
            const suffixes = ['ик', 'ор', 'ий', 'ец', 'ер'];
            
            // 1. Проверяем, есть ли хотя бы одно слово из запроса в названии
            // Если запрос пуст, то нет обязательных слов, и этот критерий не применяется.
            if (queryWords.length > 0) {
                const hasQueryWord = queryWords.some(qWord => 
                    titleWords.some(tWord => 
                        tWord === qWord || 
                        tWord.startsWith(qWord) || 
                        qWord.startsWith(tWord.slice(0, -1))
                    )
                );
                
                if (!hasQueryWord) {
                    // Если ни одно слово из запроса не найдено в названии, исключаем вакансию
                    return true; 
                }
            }

            // 2. Создаем Set корней слов из запроса для быстрой проверки
            const queryRoots = new Set();
            queryWords.forEach(w => {
                // Убираем окончания для "ик"
                if (w.endsWith('ик') || w.endsWith('ики') || w.endsWith('ика') || w.endsWith('иков') || w.endsWith('икам')) {
                    queryRoots.add(w.replace(/(ик|ики|ика|иков|икам)$/ui, ''));
                }
                // Убираем окончания для "ор"
                if (w.endsWith('ор') || w.endsWith('оры') || w.endsWith('ора') || w.endsWith('оров') || w.endsWith('орам')) {
                    queryRoots.add(w.replace(/(ор|оры|ора|оров|орам)$/ui, ''));
                }
                // Убираем окончания для "ий"
                if (w.endsWith('ий') || w.endsWith('ие') || w.endsWith('ия') || w.endsWith('их') || w.endsWith('им')) {
                    queryRoots.add(w.replace(/(ий|ие|ия|их|им)$/ui, ''));
                }
                // Убираем окончания для "ец"
                if (w.endsWith('ец') || w.endsWith('цы') || w.endsWith('ца')) {
                    queryRoots.add(w.replace(/(ец|цы|ца|цов|цам)$/ui, ''));
                }
                // Убираем окончания для "ер"
                if (w.endsWith('ер') || w.endsWith('еры') || w.endsWith('ера')) {
                    queryRoots.add(w.replace(/(ер|еры|ера|еров|ерам)$/ui, ''));
                }
                queryRoots.add(w);
            });
            
            // 3. Ищем слова с указанными суффиксами, которых нет в запросе
            for (const word of titleWords) {
                // Проверяем каждый суффикс
                for (const suffix of suffixes) {
                    if (word.endsWith(suffix) && word.length > suffix.length) {
                        const wordRoot = word.slice(0, -suffix.length);
                        
                        // Проверяем: это слово из запроса или нет?
                        const isQueryWord = queryWords.some(qw => 
                            qw === word || 
                            qw.startsWith(word) || 
                            word.startsWith(qw) ||
                            queryRoots.has(wordRoot) ||
                            qw === wordRoot
                        );
                        
                        if (!isQueryWord) {
                            excludedBySuffixDetails[suffix]++; // 📊 Считаем по типам
                            console.log(`🚫 Исключена: "${vacancyTitle}" - найдено "${word}" (суффикс "${suffix}")`);
                            return true; // Исключаем вакансию
                        }
                    }
                }
            }
            
            return false;
        }

        function startNewSearch() {
            currentPage = 0;
            hasMore = true;
            loadedCount = 0;
            withSalaryCount = 0;
            excludedCount = 0;
            excludedBySuffixDetails = { ик: 0, ор: 0, ий: 0, ец: 0, ер: 0 }; // Сбрасываем при новом поиске
            totalFound = 0;
            
            currentQuery = document.getElementById('query').value.trim() || 'склад';
            
            const tbody = document.getElementById('vacancyTableBody');
            tbody.innerHTML = ''; // Очищаем таблицу при новом поиске
            
            document.getElementById('stats').style.display = 'none';
            document.getElementById('currentSourceBadge').textContent = currentSource === 'hh' ? 'HH.ru' : 'SuperJob';
            
            allLoadedVacancies = []; // Очищаем массив при новом поиске
            loadMoreVacancies();
        }

        async function loadMoreVacancies() {
            if (isLoading || !hasMore) return;
            
            if (currentSource === 'hh') {
                await loadFromHH();
            } else {
                await loadFromSuperJob();
            }
        }

        async function loadFromHH() {
            isLoading = true;
            document.getElementById('loadingIndicator').style.display = 'block';
            
            const loadStartTime = performance.now(); // Засекаем время начала загрузки
            
            try {
                const params = new URLSearchParams({
                    text: currentQuery,
                    per_page: 10, // Загружаем по 10 вакансий
                    page: currentPage
                });

                currentCities.forEach(city => params.append('area', city));
                
                if (document.getElementById('lastTwoDays')?.checked) {
                    params.append('date_from', getDateDaysAgo(2));
                } else if (document.getElementById('freshOnly')?.checked) {
                    params.append('date_from', getDateDaysAgo(30));
                }
                if (document.getElementById('withSalary')?.checked) {
                    params.append('only_with_salary', 'true');
                }
                if (document.getElementById('noExp')?.checked) {
                    params.append('experience', 'noExperience');
                }
                if (document.getElementById('fullTime')?.checked) {
                    params.append('employment', 'full');
                }

                const response = await fetch(`https://api.hh.ru/vacancies?${params}`);
                if (!response.ok) throw new Error(`HTTP ${response.status}`);
                
                const data = await response.json();
                totalFound = data.found;
                hasMore = currentPage < data.pages - 1; // Убираем ограничение на 19 страниц

                const tbody = document.getElementById('vacancyTableBody');
                if (currentPage === 0) tbody.innerHTML = '';

                if (data.items?.length > 0) {
                    const companiesAdded = new Set();
                    const oneVacancyPerCompany = document.getElementById('oneVacancyPerCompany')?.checked;

                    for (const basicVacancy of data.items) {
                        const companyName = basicVacancy.employer?.name || 'Не указано';

                        if (oneVacancyPerCompany && companiesAdded.has(companyName)) {
                            excludedCount++;
                            continue;
                        }

                        const fullText = `${basicVacancy.name} ${companyName} ${basicVacancy.snippet?.requirement || ''} ${basicVacancy.snippet?.responsibility || ''}`.toLowerCase();
                        if (isExcluded(fullText) || shouldExcludeBySuffix(basicVacancy.name, currentQuery)) {
                            excludedCount++;
                            continue;
                        }

                        let detailedVacancy = basicVacancy;
                        let hasContacts = false; // Флаг для проверки наличия контактов

                        // Если есть HH access token, делаем запрос за детальной информацией
                        if (hhAccessToken && hhExpiresIn && Date.now() < parseInt(hhExpiresIn)) {
                            try {
                                const detailResponse = await fetch(`https://api.hh.ru/vacancies/${basicVacancy.id}`, {
                                    headers: {
                                        'Authorization': `Bearer ${hhAccessToken}`,
                                        'HH-User-Agent': 'VacancyParser/1.0 (orlov11121@mail.ru)' // Замените на ваш email
                                    }
                                });
                                if (detailResponse.ok) {
                                    detailedVacancy = await detailResponse.json();
                                    if (detailedVacancy.contacts && (detailedVacancy.contacts.phones?.length > 0 || detailedVacancy.contacts.email)) {
                                        hasContacts = true;
                                    }
                                } else {
                                    console.warn(`Failed to fetch detailed HH vacancy ${basicVacancy.id}: ${detailResponse.status}`);
                                }
                            } catch (detailError) {
                                console.error(`Error fetching detailed HH vacancy ${basicVacancy.id}:`, detailError);
                            }
                        }

                        // Фильтрация по наличию контактов
                        if (document.getElementById('onlyWithContacts')?.checked && !hasContacts) {
                            excludedCount++;
                            continue;
                        }

                        const salary = detailedVacancy.salary 
                            ? `${detailedVacancy.salary.from ? detailedVacancy.salary.from.toLocaleString() : ''}${detailedVacancy.salary.to ? ' – ' + detailedVacancy.salary.to.toLocaleString() : ''} ${detailedVacancy.salary.currency || ''}`.trim()
                            : "Не указана";

                        if (detailedVacancy.salary) withSalaryCount++;
                        loadedCount++;

                        // Форматирование контактов
                        let contactsHTML = '<span class="no-contacts">—</span>';
                        if (hasContacts) { // Используем флаг hasContacts
                            let contactParts = [];
                            
                            // Телефоны
                            if (detailedVacancy.contacts.phones && detailedVacancy.contacts.phones.length > 0) {
                                const phones = detailedVacancy.contacts.phones.map(p => {
                                    let phoneStr = p.country ? `+${p.country} ` : '';
                                    phoneStr += p.city ? `(${p.city}) ` : '';
                                    phoneStr += p.number || '';
                                    phoneStr += p.comment ? ` ${p.comment}` : '';
                                    return `<span class="contact-phone">${phoneStr}</span>`;
                                }).join('<br>');
                                contactParts.push(phones);
                            }
                            
                            // Email
                            if (detailedVacancy.contacts.email) {
                                contactParts.push(`<span class="contact-email">${detailedVacancy.contacts.email}</span>`);
                            }
                            
                            // Имя контакта
                            if (detailedVacancy.contacts.name) {
                                contactParts.push(`<span class="contact-name">${detailedVacancy.contacts.name}</span>`);
                            }
                            
                            if (contactParts.length > 0) {
                                contactsHTML = `<div class="contacts">${contactParts.join('<br>')}</div>`;
                            }
                        }

                        const row = document.createElement("tr");
                        row.innerHTML = `
                            <td class="vacancy-title">
                                ${detailedVacancy.name}
                                ${isNew(detailedVacancy.published_at) ? '<span class="badge-new">NEW</span>' : ''}
                                <span class="badge-source hh">HH</span>
                            </td>
                            <td class="company-name">${companyName}</td>
                            <td class="salary">${salary}</td>
                            <td class="location">${detailedVacancy.area?.name || 'Не указано'}</td>
                            <td class="date">${formatDate(detailedVacancy.published_at)}</td>
                            <td>${contactsHTML}</td>
                        `;
                        row.onclick = () => window.open(`https://hh.ru/vacancy/${detailedVacancy.id}`, '_blank');
                        tbody.appendChild(row);

                        if (oneVacancyPerCompany) {
                            companiesAdded.add(companyName);
                        }
                        
                        // Добавляем вакансию в общий массив для экспорта
                        allLoadedVacancies.push({
                            title: detailedVacancy.name,
                            company: companyName,
                            salary: salary,
                            location: detailedVacancy.area?.name || 'Не указано',
                            date: formatDate(detailedVacancy.published_at),
                            phone: detailedVacancy.contacts?.phones?.map(p => `${p.country ? '+' + p.country : ''}${p.city ? '(' + p.city + ')' : ''}${p.number || ''}`).join(', ') || '',
                            email: detailedVacancy.contacts?.email || '',
                            contactName: detailedVacancy.contacts?.name || ''
                        });
                    }

                    const loadEndTime = performance.now(); // Засекаем время окончания загрузки
                    const loadDuration = (loadEndTime - loadStartTime) / 1000; // Длительность в секундах
                    if (loadDuration > 0) {
                        vacanciesPerSecond = loadedCount / loadDuration;
                    }
                    
                    updateStats();
                    currentPage++;
                } else if (currentPage === 0) {
                tbody.innerHTML = `<tr><td colspan="6" class="no-results">😔 Не найдено</td></tr>`;
                hasMore = false;
            }

        } catch (error) {
            console.error('HH Error:', error);
            handleError(error);
        } finally {
            isLoading = false;
            document.getElementById('loadingIndicator').style.display = 'none';
        }
    }

        async function loadFromSuperJob() {
            isLoading = true;
            document.getElementById('loadingIndicator').style.display = 'block';
            
            if (!sjAccessToken || (sjExpiresIn && Date.now() >= parseInt(sjExpiresIn) && !(await refreshSjToken()))) {
                document.getElementById('vacancyTableBody').innerHTML = `
                    <tr><td colspan="6" style="text-align:center; padding: 40px;">
                        <h3>⚠️ SuperJob: Требуется авторизация</h3>
                        <p>Пожалуйста, войдите через SuperJob.</p>
                    </td></tr>
                `;
                isLoading = false; hasMore = false;
                document.getElementById('loadingIndicator').style.display = 'none';
                updateAuthStatus('superjob');
                return;
            }
            
            try {
                let allVacancies = [];
                let totalFoundForSJ = 0;
                let hasMoreForSJ = false;

            // If multiple cities are selected, make separate requests
            const citiesToSearch = currentCities.length > 0 ? currentCities : [0]; // Default to 'Вся Россия' if no cities selected

            for (const cityId of citiesToSearch) {
                const params = new URLSearchParams({
                    keyword: currentQuery, count: 20, page: currentPage // Загружаем по 20 вакансий
                });

                params.append('t', cityId); // Append only one city ID per request

                if (document.getElementById('lastTwoDays')?.checked) params.append('date_published_from', getUnixTimeDaysAgo(2));
                else if (document.getElementById('freshOnly')?.checked) params.append('date_published_from', getUnixTimeDaysAgo(30));
                if (document.getElementById('withSalary')?.checked) params.append('no_agreement', '1');
                if (document.getElementById('noExp')?.checked) params.append('experience', '1');
                if (document.getElementById('fullTime')?.checked) params.append('type_of_work', '6');

                const apiUrl = `https://api.superjob.ru/2.0/vacancies/?${params}`;
                const proxyUrl = `/proxy?url=${encodeURIComponent(apiUrl)}`; // Убираем key, так как токен будет в заголовке
                
                const response = await fetch(proxyUrl, {
                    headers: {
                        'Authorization': `Bearer ${sjAccessToken}`
                    }
                });
                
                if (!response.ok) {
                    const errorText = await response.text();
                    console.error('Proxy error:', errorText);
                    throw new Error(`HTTP ${response.status}: ${errorText}`);
                }
                
                const data = await response.json();
                    totalFoundForSJ += data.total;
                    if (data.more) hasMoreForSJ = true; // If any city has more, then overall has more
                    
                    if (data.objects?.length > 0) {
                        allVacancies = allVacancies.concat(data.objects);
                    }
                }

                totalFound = totalFoundForSJ;
                hasMore = hasMoreForSJ;

                const tbody = document.getElementById('vacancyTableBody');
                if (currentPage === 0) tbody.innerHTML = '';

                if (allVacancies.length > 0) {
                    const companiesAdded = new Set();
                    const oneVacancyPerCompany = document.getElementById('oneVacancyPerCompany')?.checked;

                    allVacancies.forEach(v => {
                        const companyName = v.firm_name || '—';

                        if (oneVacancyPerCompany && companiesAdded.has(companyName)) {
                            excludedCount++;
                            return;
                        }
                        
                        const fullText = `${v.profession} ${companyName} ${v.candidat || ''} ${v.work || ''}`.toLowerCase();
                        if (isExcluded(fullText) || shouldExcludeBySuffix(v.profession, currentQuery)) {
                            excludedCount++;
                            return;
                        }

                        let hasContacts = false; // Флаг для проверки наличия контактов
                        if (v.phone || v.email || v.contact) {
                            hasContacts = true;
                        }

                        // Фильтрация по наличию контактов
                        if (document.getElementById('onlyWithContacts')?.checked && !hasContacts) {
                            excludedCount++;
                            return;
                        }

                        const salary = v.payment_from || v.payment_to
                            ? `${v.payment_from ? v.payment_from.toLocaleString() : ''}${v.payment_to ? ' – ' + v.payment_to.toLocaleString() : ''} ₽`.trim()
                            : v.agreement ? "По дог." : "—";

                        if (v.payment_from || v.payment_to) withSalaryCount++;
                        loadedCount++;

                        // Форматирование контактов для SuperJob
                        let sjContactsHTML = '<span class="no-contacts">—</span>';
                        let sjContactParts = [];

                        if (hasContacts) { // Используем флаг hasContacts
                            if (v.phone) {
                                sjContactParts.push(`<span class="contact-phone">${v.phone}</span>`);
                            }
                            if (v.email) {
                                sjContactParts.push(`<span class="contact-email">${v.email}</span>`);
                            }
                            if (v.contact) {
                                sjContactParts.push(`<span class="contact-name">${v.contact}</span>`);
                            }

                            if (sjContactParts.length > 0) {
                                sjContactsHTML = `<div class="contacts">${sjContactParts.join('<br>')}</div>`;
                            }
                        }

                        const row = document.createElement("tr");
                        row.innerHTML = `
                            <td class="vacancy-title">${v.profession}${isNewFromUnix(v.date_published) ? '<span class="badge-new">NEW</span>' : ''}<span class="badge-source sj">SJ</span></td>
                            <td class="company-name">${companyName}</td>
                            <td class="salary">${salary}</td>
                            <td class="location">${v.town?.title || '—'}</td>
                            <td class="date">${formatDateFromUnix(v.date_published)}</td>
                            <td>${sjContactsHTML}</td>
                        `;
                        row.onclick = () => window.open(v.link, '_blank');
                        tbody.appendChild(row);

                        if (oneVacancyPerCompany) {
                            companiesAdded.add(companyName);
                        }

                        // Добавляем вакансию в общий массив для экспорта
                        allLoadedVacancies.push({
                            title: v.profession,
                            company: companyName,
                            salary: salary,
                            location: v.town?.title || '—',
                            date: formatDateFromUnix(v.date_published),
                            phone: v.phone || '',
                            email: v.email || '',
                            contactName: v.contact || ''
                        });
                    });

                    updateStats();
                    currentPage++;
                } else if (currentPage === 0) {
                    tbody.innerHTML = `<tr><td colspan="6" class="no-results">😔 Не найдено</td></tr>`;
                    hasMore = false;
                }

            } catch (error) {
                console.error('SJ Error:', error);
                handleError(error);
            } finally {
                isLoading = false;
                document.getElementById('loadingIndicator').style.display = 'none';
            }
        }

        function handleError(error) {
            const tbody = document.getElementById('vacancyTableBody');
            if (currentPage === 0) {
                tbody.innerHTML = `
                    <tr>
                        <td colspan="6" style="text-align:center; color:#e74c3c; padding: 40px;">
                            ⚠️ Ошибка: ${error.message}<br>
                            <button onclick="startNewSearch()" style="margin-top: 15px;">🔄 Повторить</button>
                        </td>
                    </tr>
                `;
            }
            hasMore = false;
        }

        function updateStats() {
            document.getElementById('stats').style.display = 'block';
            document.getElementById('statTotal').textContent = totalFound.toLocaleString();
            document.getElementById('statLoaded').textContent = loadedCount.toLocaleString();
            document.getElementById('statWithSalary').textContent = withSalaryCount.toLocaleString();
            document.getElementById('statExcluded').textContent = excludedCount.toLocaleString();
            
            // Обновляем статистику по суффиксам
            const suffixStat = document.getElementById('statSuffixExcluded');
            if (document.getElementById('filterBySuffix')?.checked) {
                suffixStat.style.display = 'block';
                const total = excludedBySuffixDetails.ик + excludedBySuffixDetails.ор + excludedBySuffixDetails.ий + excludedBySuffixDetails.ец + excludedBySuffixDetails.ер;
                document.getElementById('statSuffixCount').textContent = total.toLocaleString();
                document.getElementById('suffixDetails').textContent = 
                    `ик:${excludedBySuffixDetails.ик} ор:${excludedBySuffixDetails.ор} ий:${excludedBySuffixDetails.ий} ец:${excludedBySuffixDetails.ец} ер:${excludedBySuffixDetails.ер}`;
            } else {
                suffixStat.style.display = 'none';
            }

            // Обновляем статистику для вакансий с контактами (если нужно)
            // В данный момент нет отдельного счетчика для "только с контактами",
            // так как это фильтр, который уменьшает loadedCount.
            // Если нужен отдельный счетчик, его нужно добавить в логику загрузки.
        }

        window.addEventListener('scroll', () => {
            const scrollTop = window.pageYOffset || document.documentElement.scrollTop;
            const windowHeight = window.innerHeight;
            const documentHeight = document.documentElement.scrollHeight;
            
            const scrollBtn = document.getElementById('scrollTopBtn');
            scrollBtn.classList.toggle('visible', scrollTop > 300);
            
            if (scrollTop + windowHeight >= documentHeight - 500) {
                if (!isLoading && hasMore) loadMoreVacancies();
            }
        });

        function scrollToTop() {
            window.scrollTo({ top: 0, behavior: 'smooth' });
        }

        function go(platform) {
            const query = document.getElementById('query').value.trim() || "склад";
            const encoded = encodeURIComponent(query);
            const urls = {
                hh: `https://spb.hh.ru/search/vacancy?text=${encoded}`,
                avito: `https://www.avito.ru/rossiya/vakansii?q=${encoded}`,
                superjob: `https://www.superjob.ru/vacancy/search/?keywords=${encoded}`,
                zarplata: `https://www.zarplata.ru/vacancies?search=${encoded}`
            };
            if (urls[platform]) window.open(urls[platform], '_blank');
        }

        function exportToExcel() {
            if (allLoadedVacancies.length === 0) {
                alert('Нет вакансий для экспорта.');
                return;
            }

            fetch('/export_to_excel', {
                method: 'POST',
                headers: {
                    'Content-Type': 'application/json',
                },
                body: JSON.stringify({ vacancies: allLoadedVacancies }),
            })
            .then(response => {
                if (response.ok) {
                    return response.blob();
                }
                throw new Error('Ошибка экспорта в Excel');
            })
            .then(blob => {
                const url = window.URL.createObjectURL(blob);
                const a = document.createElement('a');
                a.style.display = 'none';
                a.href = url;
                a.download = 'vacancies.xlsx';
                document.body.appendChild(a);
                a.click();
                window.URL.revokeObjectURL(url);
                alert('Вакансии успешно экспортированы в Excel!');
            })
            .catch(error => {
                console.error('Ошибка при экспорте:', error);
                alert('Произошла ошибка при экспорте вакансий в Excel.');
            });
        }

        window.onload = () => {
            // Set initial default city
            const defaultId = currentSource === 'hh' ? 2 : 4;
            currentCities = [defaultId];
            
            loadCities();
            updateAuthStatus('all'); // Обновляем статус авторизации при загрузке страницы
            
            document.getElementById('exclusionInput').addEventListener('keypress', (e) => {
                if (e.key === 'Enter') addExclusion();
            });

            document.getElementById('query').addEventListener('keypress', (e) => {
                if (e.key === 'Enter') startNewSearch();
            });
        };
    </script>
</body>
</html>
'''

# ============ ЗАПУСК ============
if __name__ == '__main__':
    print('🚀 Vacancy App running!')
    print('📍 Для PythonAnywhere настройте WSGI на этот файл')
    print('📍 Для локального запуска: app.run()')
    
    # Для локального тестирования (закомментируйте на PythonAnywhere)
    app.run(host='0.0.0.0', port=8000, debug=True)
